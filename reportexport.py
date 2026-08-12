"""Reusable report exporters: Excel (.xlsx) and PDF, with the organisation letterhead.
Both take a simple (title, headers, rows) shape so any tabular report can use them."""
import io


def _org():
    import services as svc
    return svc.get_brand().get("org", "")


def _logo():
    import services as svc
    try:
        return svc.get_org_logo()   # (bytes, mime) or None
    except Exception:
        return None


def xlsx_bytes(title, headers, rows, number_cols=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    org = _org()
    wb = Workbook(); ws = wb.active; ws.title = "Report"
    navy = PatternFill("solid", fgColor="1F3864"); bw = Font(bold=True, color="FFFFFF")
    ws["A1"] = org; ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = title; ws["A2"].font = Font(bold=True, size=11, color="333333")
    r0 = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=r0, column=j, value=h); c.fill = navy; c.font = bw
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(rows, start=r0 + 1):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    for j, h in enumerate(headers, 1):
        widest = max([len(str(h))] + [len(str(row[j - 1])) for row in rows]) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(j)].width = min(42, max(10, widest + 2))
    ws.freeze_panes = f"A{r0 + 1}"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def pdf_bytes(title, headers, rows):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, Image, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    NAVY = colors.HexColor("#1f3864")
    org = _org()
    wide = len(headers) > 7
    pagesize = landscape(A4) if wide else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm, title=title)
    styles = getSampleStyleSheet()
    org_s = ParagraphStyle("org", parent=styles["Title"], textColor=NAVY, fontSize=16, alignment=0, spaceAfter=0)
    title_s = ParagraphStyle("ttl", parent=styles["Normal"], fontSize=11, textColor=colors.HexColor("#333333"), spaceAfter=6)
    cell_s = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8.5)
    head_s = ParagraphStyle("head", parent=styles["Normal"], fontSize=7.5, leading=9, textColor=colors.white, fontName="Helvetica-Bold")
    story = []
    logo = _logo()
    header_row = []
    if logo:
        try:
            data, mime = logo
            if "svg" not in (mime or ""):
                img = Image(io.BytesIO(data)); 
                ratio = img.imageHeight / float(img.imageWidth or 1)
                img.drawWidth = 42 * mm; img.drawHeight = 42 * mm * ratio
                header_row.append(img)
        except Exception:
            pass
    header_row.append(Paragraph(org, org_s))
    if header_row:
        ht = Table([header_row], colWidths=([46 * mm, None] if len(header_row) == 2 else [None]))
        ht.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
        story.append(ht)
    story.append(Paragraph(title, title_s))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=1))
    story.append(Spacer(1, 6))
    data = [[Paragraph(str(h), head_s) for h in headers]]
    for row in rows:
        data.append([Paragraph("" if v is None else str(v), cell_s) for v in row])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"
